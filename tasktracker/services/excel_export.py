"""Multi-sheet "rich workbook" export for management-style analysis, plus a
"reports bundle" CSV writer that drops one CSV per report into a folder.

Both surfaces consume the same :class:`ReportingService` shapes used by the
Reports tab so the user sees identical rows / columns whether they're
viewing in-app, opening the workbook, or attaching the bundled CSVs to
an email.

The flat ``Tasks`` / ``Todos`` / ``Notes`` / ``Activity`` sheets are pulled
directly from the ORM so they're suitable as a raw-data dump for ad-hoc
Excel pivoting (the existing single-sheet ``export_tasks_excel`` stays
intact for users who want only that).
"""

from __future__ import annotations

import csv
import datetime as dt
import html as htmllib
import re
from pathlib import Path
from typing import Any

from sqlalchemy import select
from sqlalchemy.orm import Session, selectinload

from tasktracker.db.models import (
    Task,
    TaskArea,
    TaskNote,
    TaskNoteVersion,
    TaskSubCategory,
    TaskUpdateLog,
    TodoItem,
)
from tasktracker.services.reporting_service import ReportingService, ReportResult


# ---- helpers shared by sheets ------------------------------------------------

_HEAD_RE = re.compile(r"<head\b[^>]*>.*?</head>", re.IGNORECASE | re.DOTALL)
_STYLE_RE = re.compile(r"<style\b[^>]*>.*?</style>", re.IGNORECASE | re.DOTALL)
_SCRIPT_RE = re.compile(r"<script\b[^>]*>.*?</script>", re.IGNORECASE | re.DOTALL)
_TAG_RE = re.compile(r"<[^>]+>")
_WS_RE = re.compile(r"\s+")


def _plain_from_html(value: str | None) -> str:
    """Strip HTML to plain text for the Notes / Description columns. Excel
    pivot tables are unhappy with raw HTML markup and CSV consumers can't
    do anything useful with it either."""
    if not value:
        return ""
    text = _HEAD_RE.sub(" ", value)
    text = _STYLE_RE.sub(" ", text)
    text = _SCRIPT_RE.sub(" ", text)
    text = htmllib.unescape(text)
    text = _TAG_RE.sub(" ", text)
    text = _WS_RE.sub(" ", text).strip()
    return text


def _taxonomy_labels(task: Task) -> tuple[str, str, str]:
    area = task.area
    sub = area.subcategory if area else None
    cat = sub.category if sub else None
    return (
        cat.name if cat else "",
        sub.name if sub else "",
        area.name if area else "",
    )


def _person_label(task: Task) -> tuple[str, str]:
    if task.person is None:
        return "", ""
    return f"{task.person.last_name}, {task.person.first_name}", task.person.employee_id


def _iso_date(d: dt.date | None) -> str:
    return d.isoformat() if d else ""


def _iso_datetime(ts: dt.datetime | None) -> str:
    if ts is None:
        return ""
    return ts.isoformat(sep=" ", timespec="seconds")


# ---- row builders for the flat reference sheets ------------------------------


_TASKS_HEADERS: list[str] = [
    "id",
    "ticket_number",
    "title",
    "status",
    "impact",
    "urgency",
    "priority",
    "received_date",
    "due_date",
    "closed_date",
    "next_milestone_date",
    "category",
    "subcategory",
    "area",
    "for_person",
    "for_person_employee_id",
    "age_days",
    "days_to_due",
    "days_late",
    "was_closed_late",
]


def _task_row(task: Task, as_of: dt.date) -> list[Any]:
    cat, sub, area = _taxonomy_labels(task)
    person, emp = _person_label(task)
    age_days = (as_of - task.received_date).days
    days_to_due: int | str = ""
    if task.due_date is not None:
        days_to_due = (task.due_date - as_of).days
    days_late: int | str = ""
    was_late: str = ""
    if task.due_date is not None and task.closed_date is not None:
        delta = (task.closed_date - task.due_date).days
        days_late = max(delta, 0)
        was_late = "Y" if delta > 0 else "N"
    return [
        task.id,
        task.ticket_number,
        task.title,
        task.status,
        task.impact,
        task.urgency,
        task.priority,
        _iso_date(task.received_date),
        _iso_date(task.due_date),
        _iso_date(task.closed_date),
        _iso_date(task.next_milestone_date),
        cat,
        sub,
        area,
        person,
        emp,
        age_days,
        days_to_due,
        days_late,
        was_late,
    ]


_TODOS_HEADERS: list[str] = [
    "task_id",
    "ticket_number",
    "task_title",
    "task_status",
    "todo_id",
    "sort_order",
    "todo_title",
    "milestone_date",
    "completed_at",
]


_NOTES_HEADERS: list[str] = [
    "task_id",
    "ticket_number",
    "note_id",
    "is_system",
    "created_at",
    "latest_version_seq",
    "body_text",
]


_ACTIVITY_HEADERS: list[str] = [
    "task_id",
    "ticket_number",
    "field_name",
    "old_value",
    "new_value",
    "changed_at",
]


# ---- dispatch helpers --------------------------------------------------------


def _set_table_polish(ws: Any, num_columns: int) -> None:
    """Freeze the header row and add an Excel auto-filter on the data range
    so the workbook is immediately useful without any per-sheet setup."""
    if num_columns <= 0 or ws.max_row <= 0:
        return
    ws.freeze_panes = "A2"
    last_col_letter = ws.cell(row=1, column=num_columns).column_letter
    ws.auto_filter.ref = f"A1:{last_col_letter}{ws.max_row}"


def _write_report_sheet(ws: Any, result: ReportResult) -> None:
    ws.append(result.columns)
    for row in result.rows:
        ws.append([row.get(k, "") for k in result.columns])
    _set_table_polish(ws, len(result.columns))


# ---- the rich workbook -------------------------------------------------------


def build_rich_workbook(session: Session, path: Path) -> Path:
    """Write a multi-sheet xlsx with the management reports plus the
    underlying raw data. ``path`` is created or overwritten; parent
    directories are created if missing. Returns the resolved path so
    callers can show it in a confirmation message.
    """
    from openpyxl import Workbook

    today = dt.date.today()
    rs = ReportingService(session)

    # Compute reports up front so we can build the Summary sheet first.
    wip = rs.wip_aging(as_of=today)
    weekly = rs.weekly_status(as_of=today)
    workload_r = rs.workload(as_of=today)
    first_of_month = today.replace(day=1)
    throughput_w = rs.throughput(
        from_date=first_of_month, to_date=today, period="week"
    )
    throughput_m = rs.throughput(
        from_date=first_of_month, to_date=today, period="month"
    )
    sla_r = rs.sla(from_date=first_of_month, to_date=today)
    catmix_r = rs.category_mix(from_date=first_of_month, to_date=today)

    wb = Workbook()
    # Replace default sheet so Summary lands first.
    default_sheet = wb.active
    wb.remove(default_sheet)

    # -- Summary -------------------------------------------------------------
    ws = wb.create_sheet("Summary")
    ws.append(["Task Tracker - rich workbook export"])
    ws.append(["Generated", _iso_datetime(dt.datetime.now())])
    ws.append(["As of", today.isoformat()])
    ws.append([])
    ws.append(["KPI", "Value"])
    ws.append(["Open tasks", wip.meta.get("total_open", 0)])
    for label, _lo, _hi in (("0-7", 0, 7), ("8-30", 8, 30), ("31-90", 31, 90), ("90+", 91, None)):
        ws.append([f"  Aged {label} days", wip.meta.get("buckets", {}).get(label, 0)])
    ws.append(["Closed in last 7 days", weekly.meta.get("closed_last_7", 0)])
    ws.append(["Due in next 7 days", weekly.meta.get("due_next_7", 0)])
    ws.append(["Currently blocked", weekly.meta.get("currently_blocked", 0)])
    ws.append([])
    ws.append(["SLA window", f"{first_of_month.isoformat()} -> {today.isoformat()}"])
    ws.append(["  On time", sla_r.meta.get("on_time", 0)])
    ws.append(["  Late", sla_r.meta.get("late", 0)])
    ws.append(["  Miss rate (%)", sla_r.meta.get("miss_rate_pct", 0.0)])
    ws.append(["  Avg days late", sla_r.meta.get("avg_days_late", 0.0)])
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 22

    # -- Tasks ---------------------------------------------------------------
    ws = wb.create_sheet("Tasks")
    ws.append(_TASKS_HEADERS)
    tasks_q = select(Task).options(
        selectinload(Task.area)
        .selectinload(TaskArea.subcategory)
        .selectinload(TaskSubCategory.category),
        selectinload(Task.person),
    )
    tasks_all = list(session.scalars(tasks_q).unique().all())
    tasks_all.sort(key=lambda t: (t.ticket_number is None, t.ticket_number or 0))
    for t in tasks_all:
        ws.append(_task_row(t, today))
    _set_table_polish(ws, len(_TASKS_HEADERS))

    # -- Todos ---------------------------------------------------------------
    ws = wb.create_sheet("Todos")
    ws.append(_TODOS_HEADERS)
    todo_q = (
        select(TodoItem)
        .options(selectinload(TodoItem.task))
        .order_by(TodoItem.task_id, TodoItem.sort_order)
    )
    for todo in session.scalars(todo_q).all():
        task = todo.task
        ws.append(
            [
                task.id,
                task.ticket_number,
                task.title,
                task.status,
                todo.id,
                todo.sort_order,
                todo.title,
                _iso_date(todo.milestone_date),
                _iso_datetime(todo.completed_at),
            ]
        )
    _set_table_polish(ws, len(_TODOS_HEADERS))

    # -- Notes ---------------------------------------------------------------
    ws = wb.create_sheet("Notes")
    ws.append(_NOTES_HEADERS)
    notes_q = (
        select(TaskNote)
        .options(
            selectinload(TaskNote.task),
            selectinload(TaskNote.versions),
        )
        .order_by(TaskNote.task_id, TaskNote.created_at)
    )
    for note in session.scalars(notes_q).unique().all():
        latest: TaskNoteVersion | None = (
            max(note.versions, key=lambda v: v.version_seq) if note.versions else None
        )
        ws.append(
            [
                note.task_id,
                note.task.ticket_number if note.task else "",
                note.id,
                "Y" if note.is_system else "N",
                _iso_datetime(note.created_at),
                latest.version_seq if latest else "",
                _plain_from_html(latest.body_html if latest else ""),
            ]
        )
    _set_table_polish(ws, len(_NOTES_HEADERS))

    # -- Activity ------------------------------------------------------------
    ws = wb.create_sheet("Activity")
    ws.append(_ACTIVITY_HEADERS)
    log_q = (
        select(TaskUpdateLog)
        .options(selectinload(TaskUpdateLog.task))
        .order_by(TaskUpdateLog.task_id, TaskUpdateLog.changed_at)
    )
    for entry in session.scalars(log_q).all():
        ws.append(
            [
                entry.task_id,
                entry.task.ticket_number if entry.task else "",
                entry.field_name,
                entry.old_value or "",
                entry.new_value or "",
                _iso_datetime(entry.changed_at),
            ]
        )
    _set_table_polish(ws, len(_ACTIVITY_HEADERS))

    # -- Report sheets -------------------------------------------------------
    _write_report_sheet(wb.create_sheet("WIP_Aging"), wip)
    _write_report_sheet(wb.create_sheet("Throughput_Weekly"), throughput_w)
    _write_report_sheet(wb.create_sheet("Throughput_Monthly"), throughput_m)
    _write_report_sheet(wb.create_sheet("Workload"), workload_r)
    _write_report_sheet(wb.create_sheet("SLA"), sla_r)
    _write_report_sheet(wb.create_sheet("CategoryMix"), catmix_r)
    _write_report_sheet(wb.create_sheet("WeeklyStatus"), weekly)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


# ---- per-report CSV bundle ---------------------------------------------------


def write_reports_bundle_csvs(session: Session, folder: Path) -> list[Path]:
    """Write one CSV per report into ``folder`` and return the list of
    written paths. The ``folder`` is created if missing. Files are named
    after the report id (``wip_aging.csv`` etc.) so an emailed bundle is
    self-explanatory."""
    folder.mkdir(parents=True, exist_ok=True)
    today = dt.date.today()
    first_of_month = today.replace(day=1)
    rs = ReportingService(session)

    bundle: list[tuple[str, ReportResult]] = [
        ("wip_aging", rs.wip_aging(as_of=today)),
        ("throughput_weekly", rs.throughput(first_of_month, today, period="week")),
        ("throughput_monthly", rs.throughput(first_of_month, today, period="month")),
        ("workload", rs.workload(as_of=today)),
        ("sla", rs.sla(first_of_month, today)),
        ("category_mix", rs.category_mix(first_of_month, today)),
        ("weekly_status", rs.weekly_status(as_of=today)),
    ]

    written: list[Path] = []
    for slug, result in bundle:
        out = folder / f"{slug}.csv"
        with out.open("w", newline="", encoding="utf-8") as fh:
            writer = csv.DictWriter(fh, fieldnames=result.columns)
            writer.writeheader()
            for row in result.rows:
                writer.writerow({k: row.get(k, "") for k in result.columns})
        written.append(out)
    return written
